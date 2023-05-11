from django.shortcuts import render, redirect
from rjd.form.form import *
from app_cw.utils.pagination import Pagination
import datetime
from openpyxl import load_workbook
import re
from rjd.utils.excledate import *


def account_pub_list(request):
    """对公已付表"""
    now_time = datetime.datetime.now()
    data_dict = {}
    # 数据收集
    search_data = request.GET.get('q', "")
    search_date = request.GET.get('d', '')
    search_date2 = request.GET.get('d2', '')
    search_month = request.GET.get('month', '')
    search_year = request.GET.get('year', '')
    if not search_year:
        search_year = now_time.year - 1
    if search_month:
        queryset = models.PublicAccount.objects.filter(create_time__year=search_year,
                                                       create_time__month=search_month).order_by('-id')
    else:
        if search_date:
            start_date = datetime.date(*map(int, search_date.split('-')))
        else:
            start_date = ''
        if search_date2:
            end_date = datetime.date(*map(int, search_date2.split('-')))
        else:
            end_date = ''
        if search_data:
            data_dict["supplier__supplier_name__contains"] = search_data
        if start_date and end_date:
            queryset = models.PublicAccount.objects.filter(**data_dict,
                                                           create_time__range=(start_date, end_date)).order_by('-id')
        else:
            # queryset = models.PublicAccount.objects.filter(**data_dict, create_time__year=now_time.year,
            #                                                create_time__month=now_time.month).order_by('-id')
            queryset = models.PublicAccount.objects.filter(**data_dict).order_by('-id')

    page_object = Pagination(request, queryset)
    context = {
        'queryset': page_object.page_queryset,
        'search_data': search_data,
        'page_string': page_object.html(),
        'search_date': search_date,
        'search_date2': search_date2,
        'search_month':search_month,
        'search_year':search_year,
    }
    return render(request, 'account_pub_list.html', context)


def account_pub_add(request):
    """添加已付账户"""
    if request.method == "GET":
        form = AccountPubModelForm()
        context = {
            'form': form,
            'title': '添加对公付款'
        }
        return render(request, 'rjd_add.html', context)
    form = AccountPubModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/rjd/account/pub/list/')
    context = {
        'form': form,
        'title': "添加对公付款",
    }
    return render(request, 'rjd_add.html', context)


def account_pub_edit(request, nid):
    """编辑已付账单"""
    row_object = models.PublicAccount.objects.filter(id=nid).first()

    if request.method == "GET":
        form = AccountPubModelForm(instance=row_object)
        context = {
            'form': form,
            'title': '编辑对公已付',
        }
        return render(request, 'rjd_add.html', context)
    form = AccountPubModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/rjd/account/pub/list/')
    context = {
        'form': form,
        'title': '编辑对公已付',
    }
    return render(request, 'rjd_add.html', context)


def account_pub_delete(request, nid):
    """删除单行账目"""
    models.PublicAccount.objects.filter(id=nid).delete()
    return redirect('/rjd/account/pub/list/')


def account_pub_multi(request):
    """批量上传对公账户"""
    # 创建空的对象列表
    user_list = []
    # 1、获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    # 2、对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    for sheet in wb.worksheets:
        month = re.sub(r'[^0-9]', "", sheet.title)  # 字符串替换字体
        print("________________%s____________________" % (month))
        for row in sheet.iter_rows(min_row=8):
            if row[0].value == None:
                continue
            else:
                t = xldate_as_datetime(row[1].value)  # datetime格式时间
                user_list.append(models.PublicAccount(
                    create_time=t, supplier=row[2].value, product_name=row[3].value, product_model=row[4].value,
                    unit_price=row[7].value, total_price=row[8].value, tax=row[10].value, bill=row[11].value,
                ))
    # print(len(user_list))
    models.PublicAccount.objects.bulk_create(user_list)
    return redirect('/rjd/account/pub/list/')
